#!/usr/bin/python3

# Main file for track session analysis.

from dataImporter import *
from datamodel import TrackSession
import pathlib
import argparse
from openpyxl import load_workbook
import datetime
import os, os.path, shutil, io
import utils
import base64

parser = argparse.ArgumentParser(description='Update lap time record spreadsheet')
parser.add_argument('-d', '--dir', action='append', help='Directory containing CSV datafiles with lap data to be recorded.')
parser.add_argument('-v', '--verbose', action='count')
parser.add_argument('-t', '--trackname', action='store', help='Name of track data is from, if not present in file.')
parser.add_argument('--gps-only', action='store_true', help='Perform analysis only on GPS data (e.g. AIM Solo 2 non-DL data)', default=False)

args = parser.parse_args()

def debugout(debuglevel, text):
    if args.verbose and args.verbose >= debuglevel:
        print(text)

def analyze(dirName, sessions):
    sessionTimes = [s.getSessionInfo("sessionTime") for s in sessions]

#    outputFilename = '-'.join([session.getSessionInfo("driverName"),
#                               session.getSessionInfo("trackName"),
#                               session.getSessionInfo("sessionDate"),
#                               session.getSessionInfo("sessionTime")])
#    outputFilename += '.pdf'

    # Tell us which session we're looking at
    #debugout(1, f'Analyzing data from { sessions.getSessionInfo("sourcefile") }')
    #if session.getSessionInfo('trackName') is not None:
    #    debugout(1, f'Track name: { session.getSessionInfo("trackName") }') 
    #if session.getSessionInfo('trackDescription') is not None:
    #    debugout(1, f'Track description: { session.getSessionInfo("trackDescription") }') 
    #if session.getSessionInfo('sessionDate') is not None:
    #    debugout(1, f'Session date: { session.getSessionInfo("sessionDate") }') 
    #if session.getSessionInfo('sessionTime') is not None:
    #    debugout(1, f'Session time: { session.getSessionInfo("sessionTime") }') 

    debugout(2, "Opening workbook")
    workbook = load_workbook("/nfshome/jberning/TrackTimes.xlsx")
    if dirName not in workbook.sheetnames:
        debugout(1, "Adding sheet name "+dirName)
        sheet = workbook.create_sheet(title=dirName)
    else:
        sheet = workbook[dirName]

    sheet['A1'] = "Date"
    sheet['B1'] = sessions[0].getSessionInfo("sessionDate")
    sheet['A2'] = "Track"
    sheet['B2'] = sessions[0].getSessionInfo("trackName")
    sheet['A6'] = "Lap#"
    sheet['A25'] = "Std Dev"
    for f in range(15):
        sheet['A'+str(7+f)] = f+1
    # Set time heading, fill in the times
    for idx, session in enumerate(sessions):
        column = chr(ord('B')+idx)
        sheet[column+'6'] = session.getSessionInfo("sheetDateTime")
        for lidx, lapTime in enumerate(session.getLapTimes()[1:-1]):
            sheet[column+str(7+lidx)] = datetime.timedelta(seconds=lapTime)
            sheet[column+str(7+lidx)].number_format = "mm:ss.000"
        sheet[column+'25'] = '=_xlfn.STDEV.P('+column+'7:'+column+'21)'
        sheet[column+'25'].number_format = "mm:ss.000"
    
    sheet['A28'] = "Fastest overall lap:"
    sheet['A29'] = "=MIN(B7:I21)"
    sheet['A29'].number_format = "mm:ss.000"
    for col in 'BCDEFGHI':
        sheet.column_dimensions[col].width = 15

    workbook.save("/nfshome/jberning/TrackTimes.xlsx")

def slurpDir(dirName):
    if dirName.endswith('/'):
        dirName = dirName[:-1]
    files = [str(f) for f in pathlib.Path().glob(dirName+"/*.csv")]

    runs = []
    for file in files:
        debugout(1, "Working on file: "+file)
        dataReader = getFileImporter(file)
        runs.append(dataReader.readSessionData(args))
        runs[-1].addSessionInfo(sourcefile = file)


    analyze(os.path.basename(dirName), runs)

def main():
    for dir in args.dir:
        slurpDir(dir)

if __name__ == '__main__':
    main()
